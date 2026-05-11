import streamlit as st
import pandas as pd
import re
import math
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Parámetros ─────────────────────────────────────────────
ANELO_LAT      = -38.3500
ANELO_LON      = -68.7833
RADIO_KM       = 15
NQN_LAT_NORTE  = -38.830
NQN_LAT_SUR    = -38.990
NQN_LON_OESTE  = -68.155
NQN_LON_ESTE   = -68.015

NOMBRES_ANELO = [
    "añelo","anelo","parque industrial añelo","parque industrial anelo",
    "paruqe industrial añelo","bajo añelo","bajo anelo",
    "bajada de añelo","bajada de anelo","bajada añelo","ban","tratayen","loma campana",
]
NOMBRES_NEUQUEN = [
    "parque industrial neuquen","parque industrial neuquén","pin neuquen","pin neuquén","pin nqn",
    "zona 1","z1","zona industrial",
    "carlos pellegrini","enrique mosconi","gobernador belenguer",
    "trabajadores de la industria","solalique","entre rios",
    "expedicionarios del desierto","isla victoria","dewey",
    "juan julian lastra","dr ramon","dr. ramon","hipolito yrigoyen",
    "tte coronel","tte. coronel","tte de navio","tte. de navio",
    "antartida argentina","av argentina","av. argentina",
    "mariano moreno","industriales neuquinos","emilio belenguer","lote 12 mza",
    "makro neuquen","la anonima shopping","hotel del comahue","colegio lincoln",
    "torre trade","edificio zafiro","edificio dye","edificio renova","centro oncologico",
    "domuyo centro de convenciones","predio fnc","base swpetrol","base pataco",
    "base manpetrol neuquen","hormigonera del interior pin","embotelladora del atlantico",
    "almacen pecom pin","base pecom pin","oficinas tgs pin","planta de asfalto omega",
    "centro de logistica alberta","obrador centro de capacitaciones","logistica la serenisima",
    "technip fmc","polyar","segar","tiser",
]

def haversine(la1,lo1,la2,lo2):
    R=6371.0; dlat=math.radians(la2-la1); dlon=math.radians(lo2-lo1)
    a=math.sin(dlat/2)**2+math.cos(math.radians(la1))*math.cos(math.radians(la2))*math.sin(dlon/2)**2
    return R*2*math.asin(math.sqrt(a))

def esta_en_ejido_neuquen(lat,lon):
    return NQN_LAT_NORTE >= lat >= NQN_LAT_SUR and NQN_LON_OESTE <= lon <= NQN_LON_ESTE

def extraer_coords(texto):
    texto=str(texto)
    m=re.search(r'(-(?:3[5-9]|4[0-3])\.\d+)[,\s]+(-(?:6[6-9]|7[0-2])\.\d+)',texto)
    if m:
        lat,lon=float(m.group(1)),float(m.group(2))
        if -43<lat<-35 and -72<lon<-66: return lat,lon
    m=re.search(r'(\d{2})°(\d{2})\'([\d.]+)[\"″\u201d][Ss]\s+(\d{2,3})°(\d{2})\'([\d.]+)[\"″\u201d][Ww]',texto)
    if m:
        lat=-(int(m.group(1))+int(m.group(2))/60+float(m.group(3))/3600)
        lon=-(int(m.group(4))+int(m.group(5))/60+float(m.group(6))/3600)
        return lat,lon
    return None

def normalizar(texto):
    if not isinstance(texto,str): return ""
    r=str.maketrans("áéíóúàèìòùäëïöüñÁÉÍÓÚÑ","aeiouaeiouaeiounAEIOUN")
    return texto.lower().strip().translate(r)

def extraer_nombre_localidad(texto_raw):
    texto=str(texto_raw)
    texto=re.split(r'Neuquen[-–]',texto,flags=re.IGNORECASE)[0]
    texto=re.sub(r'-?\d{2,3}\.\d+[,\s]*',' ',texto)
    texto=re.sub(r'\d{1,3}°\d{1,2}\'[\d.]+[\"″\u201d]?[SsWwNnEe]?',' ',texto)
    texto=re.sub(r'https?://\S+',' ',texto)
    texto=re.sub(r'[,;\(\)\[\]|]+',' ',texto)
    return re.sub(r'\s+',' ',texto).strip()

def clasificar(texto_raw):
    if not texto_raw or str(texto_raw).strip() in ["","-",".","-  -","Lug Trabajo","X2",". ."]:
        return "","Sin información"
    coords=extraer_coords(texto_raw)
    if coords:
        lat,lon=coords
        dist=haversine(ANELO_LAT,ANELO_LON,lat,lon)
        if dist<=RADIO_KM: return "Añelo",f"{dist:.1f} km de Añelo"
        if esta_en_ejido_neuquen(lat,lon): return "Neuquén",f"Ejido Nqn Capital"
        return "Otros",f"{dist:.1f} km de Añelo"
    localidad=normalizar(extraer_nombre_localidad(texto_raw))
    for n in NOMBRES_ANELO:
        if re.search(r'\b'+re.escape(normalizar(n))+r'\b',localidad): return "Añelo","Sin GPS - nombre"
    for n in NOMBRES_NEUQUEN:
        if normalizar(n) in localidad: return "Neuquén","Sin GPS - nombre"
    return "Otros","Sin GPS"

def procesar_excel(uploaded_file):
    df_raw=pd.read_excel(uploaded_file,header=None)
    fila_header=0
    for i,row in df_raw.iterrows():
        if any("lug trabajo" in str(v).lower() for v in row.values if v):
            fila_header=i; break
    uploaded_file.seek(0)
    df=pd.read_excel(uploaded_file,header=fila_header)
    df.columns=[str(c).strip() for c in df.columns]
    col_lugar=df.columns[1]; col_anelo=df.columns[2]
    col_neuquen=df.columns[3]; col_otro=df.columns[4]
    resultados=df[col_lugar].apply(clasificar)
    df["_clasif"]=resultados.apply(lambda x:x[0])
    df["Observaciones"]=resultados.apply(lambda x:x[1])
    df[col_anelo]  =df["_clasif"].apply(lambda x:"X" if x=="Añelo"         else "")
    df[col_neuquen]=df["_clasif"].apply(lambda x:"X" if x=="Neuquén"        else "")
    df[col_otro]   =df["_clasif"].apply(lambda x:"X" if x=="Otros"  else "")
    df.rename(columns={col_neuquen:"Neuquén",col_otro:"Otros"},inplace=True)
    df.drop(columns=["_clasif"],inplace=True)
    cant_a=(df[col_anelo]=="X").sum()
    cant_n=(df["Neuquén"]=="X").sum()
    cant_o=(df["Otros"]=="X").sum()
    total=len(df)
    output=io.BytesIO(); df.to_excel(output,index=False); output.seek(0)
    wb=load_workbook(output); ws=wb.active
    C_HEADER="1F4E79"; C_ANELO="FFF2CC"; C_NQN="E2EFDA"; C_OTROS="DAE3F3"
    fill_hdr=PatternFill("solid",fgColor=C_HEADER)
    font_hdr=Font(bold=True,color="FFFFFF",name="Arial",size=10)
    font_nm=Font(name="Arial",size=10)
    font_xa=Font(bold=True,color="7F6000",name="Arial",size=11)
    font_xn=Font(bold=True,color="375623",name="Arial",size=11)
    font_xo=Font(bold=True,color="1F4E79",name="Arial",size=11)
    a_c=Alignment(horizontal="center",vertical="center")
    a_l=Alignment(horizontal="left",vertical="center",wrap_text=True)
    thin=Side(style="thin",color="CCCCCC"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    for col in range(1,ws.max_column+1):
        c=ws.cell(row=1,column=col); c.fill=fill_hdr; c.font=font_hdr; c.alignment=a_c; c.border=brd
    for row in range(2,ws.max_row+1):
        for col in range(1,ws.max_column+1):
            cell=ws.cell(row=row,column=col); cell.font=font_nm; cell.border=brd; cell.alignment=a_l
            if col==3 and cell.value=="X":
                cell.alignment=a_c; cell.fill=PatternFill("solid",fgColor=C_ANELO); cell.font=font_xa
            elif col==4 and cell.value=="X":
                cell.alignment=a_c; cell.fill=PatternFill("solid",fgColor=C_NQN);   cell.font=font_xn
            elif col==5 and cell.value=="X":
                cell.alignment=a_c; cell.fill=PatternFill("solid",fgColor=C_OTROS);  cell.font=font_xo
            if col==ws.max_column and cell.value:
                cell.font=Font(name="Arial",size=9,italic=True,color="595959")
                cell.fill=PatternFill("solid",fgColor="F2F2F2")
    anchos={1:22,2:60,3:10,4:14,5:16,6:14,7:40,8:35}
    for c,w in anchos.items():
        if c<=ws.max_column: ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    final=io.BytesIO(); wb.save(final); final.seek(0)
    return final,cant_a,cant_n,cant_o,total

# ── UI ─────────────────────────────────────────────────────
st.set_page_config(page_title="Clasificador Tributario · Sullair",page_icon="⚡",layout="centered")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html,body,[class*="css"]{font-family:'IBM Plex Sans',sans-serif;}
.block-container{max-width:720px;padding-top:2rem;}
h1{font-family:'IBM Plex Mono',monospace;font-size:1.6rem;color:#1F4E79;letter-spacing:-0.5px;
   border-bottom:3px solid #1F4E79;padding-bottom:0.4rem;margin-bottom:0.2rem;}
.subtitulo{font-size:0.85rem;color:#666;margin-bottom:2rem;font-family:'IBM Plex Mono',monospace;}
.regla-box{background:#f0f4f9;border-left:4px solid #1F4E79;padding:0.8rem 1.2rem;
   border-radius:0 6px 6px 0;font-size:0.84rem;color:#333;margin-bottom:1.5rem;line-height:1.7;}
.metric-row{display:flex;gap:0.8rem;margin:1.5rem 0;}
.metric-card{flex:1;padding:1rem;border-radius:8px;text-align:center;}
.metric-card.anelo{background:#FFF2CC;border:1px solid #f0d060;}
.metric-card.nqn{background:#E2EFDA;border:1px solid #9dc77a;}
.metric-card.otros{background:#DAE3F3;border:1px solid #a0b8dc;}
.metric-card .numero{font-family:'IBM Plex Mono',monospace;font-size:1.9rem;font-weight:600;line-height:1;}
.metric-card.anelo .numero{color:#7F6000;}
.metric-card.nqn .numero{color:#375623;}
.metric-card.otros .numero{color:#1F4E79;}
.metric-card .label{font-size:0.78rem;color:#555;margin-top:0.3rem;}
</style>
""",unsafe_allow_html=True)

st.markdown("<h1>⚡ Clasificador Tributario</h1>",unsafe_allow_html=True)
st.markdown('<p class="subtitulo">Sullair · Alquileres Oil & Gas · Neuquén</p>',unsafe_allow_html=True)
st.markdown("""
<div class="regla-box">
<b>Reglas de clasificación:</b><br>
🟡 <b>Añelo</b> — GPS ≤ 15 km del centro de Añelo &nbsp;|&nbsp; Sin GPS: Añelo, Bajo Añelo, Bajada de Añelo, BAN, Tratayen, Loma Campana<br>
🟢 <b>Neuquén</b> — GPS dentro del ejido municipal (límites: Centenario al N · Río Limay al S · Plottier al O · Río Neuquén al E) &nbsp;|&nbsp; Sin GPS: barrios, PIN, instalaciones de la ciudad<br>
🔵 <b>Otros</b> — Todo lo que no entra en las categorías anteriores
</div>
""",unsafe_allow_html=True)

archivo=st.file_uploader("Cargá el archivo Excel de facturas",type=["xlsx"],
    help="El archivo debe tener: Nº Factura | Lug. Trabajo | Añelo | Ciudad de Neuquén | Otro")

if archivo:
    with st.spinner("Clasificando facturas..."):
        try:
            resultado,cant_a,cant_n,cant_o,total=procesar_excel(archivo)
            st.success(f"✅ {total:,} facturas procesadas")
            pa=cant_a/total*100 if total else 0
            pn=cant_n/total*100 if total else 0
            po=cant_o/total*100 if total else 0
            st.markdown(f"""
            <div class="metric-row">
                <div class="metric-card anelo">
                    <div class="numero">{cant_a:,}</div>
                    <div class="label">Añelo &nbsp;·&nbsp; {pa:.1f}%</div>
                </div>
                <div class="metric-card nqn">
                    <div class="numero">{cant_n:,}</div>
                    <div class="label">Neuquén &nbsp;·&nbsp; {pn:.1f}%</div>
                </div>
                <div class="metric-card otros">
                    <div class="numero">{cant_o:,}</div>
                    <div class="label">Otros &nbsp;·&nbsp; {po:.1f}%</div>
                </div>
            </div>
            """,unsafe_allow_html=True)
            st.download_button(
                label="⬇️  Descargar Excel clasificado",
                data=resultado,
                file_name="facturas_clasificadas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")
            st.info("Verificá que el archivo tenga el formato correcto (columnas: Nº Factura, Lug. Trabajo, Añelo, Ciudad de Neuquén, Otro).")
